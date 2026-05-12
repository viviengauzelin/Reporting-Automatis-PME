"""
dashboard.py - Classeur Excel enrichi : tableau de bord visuel et rapport de réconciliation.

Module additionnel V2.1 — aucune fonction de utils.py n'est modifiée.

Nouvelles fonctions publiques :
    compute_dashboard_kpis()              → indicateurs clés pré-calculés
    build_excel_bytes_with_dashboard()    → export mode Streamlit
    export_excel_with_dashboard()         → export mode Batch

Structure du classeur produit (5 feuilles dans l'ordre) :
    1. 📊 Dashboard      → KPIs + 5 graphiques
    2. 📅 Par mois       → agrégation mensuelle formatée
    3. 👤 Par commercial → agrégation par commercial formatée
    4. 🔍 Données brutes → transactions nettoyées
    5. 🔐 Réconciliation → rapport d'intégrité (miroir du log .txt)
    (+ _data, feuille cachée alimentant les graphiques)

Palette : bleu marine (#1F3864), gris, blanc — sobre, adapté direction financière PME.
Graphiques : 5 graphiques matplotlib (PNG) — titres et axes positionnés en dehors de la zone
    de tracé, graduations complètes avec formatage monétaire.
"""

from __future__ import annotations

import io
import logging
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
plt.switch_backend("agg")   # backend non-interactif — requis en environnement serveur (Streamlit)
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import SETTINGS
from utils import ReconciliationReport

# ===========================================================================
# PALETTE DE COULEURS — Charte sobre (bleu marine / gris / blanc)
# ===========================================================================
# Hexadécimal sans '#', format openpyxl.

_NAVY       = "1F3864"   # bleu marine foncé  — titres, valeurs KPI
_NAVY_MID   = "2E75B6"   # bleu moyen         — accents, courbes secondaires
_GREY_DARK  = "404040"   # gris foncé         — texte courant
_GREY_MID   = "7F7F7F"   # gris moyen         — libellés secondaires
_GREY_LITE  = "F2F2F2"   # gris très clair    — fond cartes KPI, alternance lignes
_BORDER_COL = "D9D9D9"   # gris clair         — bordures
_WHITE      = "FFFFFF"   # blanc              — fond principal
_GREEN      = "375623"   # vert foncé sobre   — variation positive
_RED        = "C00000"   # rouge foncé sobre  — variation négative / alerte
_ALERT_BG   = "FFF2CC"   # jaune très pâle    — fond bloc alerte réconciliation

# Dimensions des graphiques (cm)
_CHART_W = 13.5
_CHART_H = 10.5

# Largeurs des colonnes de la feuille Dashboard
_COL_WIDTHS_DB: dict[int, float] = {
    1: 20,   # A — KPI col 1
    2: 3,    # B — espaceur
    3: 20,   # C — KPI col 2
    4: 3,    # D — espaceur
    5: 20,   # E — KPI col 3
    6: 3,    # F — espaceur droit
}

# ===========================================================================
# STYLES RÉUTILISABLES
# ===========================================================================

_THIN_SIDE   = Side(border_style="thin", color=_BORDER_COL)
_THIN_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE,
    top=_THIN_SIDE, bottom=_THIN_SIDE,
)


def _font(
    bold: bool = False,
    size: int = 10,
    color: str = _GREY_DARK,
    italic: bool = False,
) -> Font:
    """Construit un objet Font Calibri avec les paramètres donnés."""
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)


def _fill(color: str) -> PatternFill:
    """Construit un PatternFill uni."""
    return PatternFill(fill_type="solid", fgColor=color)


# ===========================================================================
# HELPER D'ÉCRITURE DE CELLULE
# ===========================================================================

def _w(
    ws: Any,
    row: int,
    col: int,
    value: Any = None,
    *,
    bold: bool = False,
    size: int = 10,
    color: str = _GREY_DARK,
    italic: bool = False,
    bg: Optional[str] = None,
    align: str = "left",
    wrap: bool = False,
    num_fmt: str = "General",
) -> Any:
    """Écrit une valeur dans une cellule et applique le style demandé.

    Fonction volontairement courte (nommée ``_w``) car appelée très souvent
    dans les builders de feuilles. Ne gère pas les bordures, qui sont appliquées
    manuellement quand un contrôle précis par côté est nécessaire.

    Args:
        ws: Feuille openpyxl cible.
        row: Ligne (1-indexé).
        col: Colonne (1-indexé).
        value: Valeur à écrire (None = ne pas modifier la valeur).
        bold: Gras.
        size: Taille de police.
        color: Couleur hex de la police.
        italic: Italique.
        bg: Couleur hex de fond (None = pas de fond).
        align: Alignement horizontal ("left", "center", "right").
        wrap: Retour à la ligne automatique.
        num_fmt: Format numérique Excel (ex: ``'#,##0.00 "€"'``).

    Returns:
        La cellule modifiée.
    """
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    cell.font      = _font(bold=bold, size=size, color=color, italic=italic)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg:
        cell.fill = _fill(bg)
    if num_fmt != "General":
        cell.number_format = num_fmt
    return cell


def _fmt_eur(x: float, decimals: int = 0) -> str:
    """Formate un montant en euros avec séparateur de milliers français.

    Args:
        x: Montant numérique.
        decimals: Nombre de décimales (0 ou 2).

    Returns:
        Chaîne formatée, ex: ``"1 234 567 €"`` ou ``"1 234,56 €"``.
    """
    if decimals == 0:
        return f"{x:,.0f} €".replace(",", " ")
    s = f"{x:,.2f}".replace(",", "X").replace(".", ",").replace("X", " ")
    return f"{s} €"


# ===========================================================================
# DATACLASS — Indicateurs clés du tableau de bord
# ===========================================================================

@dataclass
class DashboardKPIs:
    """Indicateurs clés pré-calculés pour le tableau de bord DAF.

    Attributes:
        total_revenue: CA total de la période (€).
        transaction_count: Nombre de transactions avec montant valide.
        average_basket: Panier moyen en € (CA / nb transactions valides).
        monthly_avg_revenue: CA moyen par mois.
        active_months: Nombre de mois distincts couverts.
        best_month: Mois YYYY-MM au CA le plus élevé.
        best_month_revenue: CA du meilleur mois.
        worst_month: Mois YYYY-MM au CA le plus faible.
        worst_month_revenue: CA du pire mois.
        last_mom_growth_pct: Variation M/M % entre les 2 derniers mois.
            None si données insuffisantes (< 2 mois).
        top_salesperson: Commercial avec le plus haut CA. None si absent.
        top_salesperson_revenue: CA du top commercial.
        active_salespeople: Nombre de commerciaux distincts.
    """

    total_revenue: float
    transaction_count: int
    average_basket: float
    monthly_avg_revenue: float
    active_months: int
    best_month: str
    best_month_revenue: float
    worst_month: str
    worst_month_revenue: float
    last_mom_growth_pct: Optional[float]
    top_salesperson: Optional[str]
    top_salesperson_revenue: Optional[float]
    active_salespeople: int


def compute_dashboard_kpis(
    df: pd.DataFrame,
    report_months: pd.DataFrame,
    report_salespeople: pd.DataFrame,
) -> DashboardKPIs:
    """Calcule les indicateurs clés à partir des données nettoyées.

    Tous les calculs sont basés sur les montants **valides** (non-NaN) uniquement,
    conformément à la règle comptable : un montant non convertible n'est pas du CA.

    Args:
        df: DataFrame nettoyé (colonnes : date, montant, mois, commercial optionnel).
        report_months: Agrégation mensuelle [mois, montant], non triée.
        report_salespeople: Agrégation par commercial [commercial, montant] ou vide.

    Returns:
        DashboardKPIs avec tous les indicateurs calculés.
    """
    valid_amounts     = df["montant"].dropna()
    total_revenue     = float(valid_amounts.sum())
    transaction_count = int(valid_amounts.count())
    average_basket    = total_revenue / transaction_count if transaction_count > 0 else 0.0

    months_sorted     = report_months.sort_values("mois").reset_index(drop=True)
    active_months     = int(months_sorted["mois"].nunique())
    monthly_avg       = total_revenue / active_months if active_months > 0 else 0.0

    idx_best  = int(months_sorted["montant"].idxmax())
    idx_worst = int(months_sorted["montant"].idxmin())
    best_month          = str(months_sorted.loc[idx_best,  "mois"])
    best_month_revenue  = float(months_sorted.loc[idx_best,  "montant"])
    worst_month         = str(months_sorted.loc[idx_worst, "mois"])
    worst_month_revenue = float(months_sorted.loc[idx_worst, "montant"])

    # Variation M/M : compare les 2 derniers mois connus.
    # Justification métier : on ne calcule pas sur la moyenne pour éviter
    # de noyer une rupture récente dans l'historique.
    last_mom_growth_pct: Optional[float] = None
    if len(months_sorted) >= 2:
        last_val = float(months_sorted["montant"].iloc[-1])
        prev_val = float(months_sorted["montant"].iloc[-2])
        if prev_val != 0:
            last_mom_growth_pct = (last_val - prev_val) / abs(prev_val) * 100.0

    top_salesperson:         Optional[str]   = None
    top_salesperson_revenue: Optional[float] = None
    active_salespeople = 0

    if not report_salespeople.empty and len(report_salespeople) > 0:
        active_salespeople = int(report_salespeople["commercial"].nunique())
        idx_top = int(report_salespeople["montant"].idxmax())
        top_salesperson         = str(report_salespeople.loc[idx_top, "commercial"])
        top_salesperson_revenue = float(report_salespeople.loc[idx_top, "montant"])

    return DashboardKPIs(
        total_revenue=total_revenue,
        transaction_count=transaction_count,
        average_basket=average_basket,
        monthly_avg_revenue=monthly_avg,
        active_months=active_months,
        best_month=best_month,
        best_month_revenue=best_month_revenue,
        worst_month=worst_month,
        worst_month_revenue=worst_month_revenue,
        last_mom_growth_pct=last_mom_growth_pct,
        top_salesperson=top_salesperson,
        top_salesperson_revenue=top_salesperson_revenue,
        active_salespeople=active_salespeople,
    )



# ===========================================================================
# 5 GRAPHIQUES — Builders individuels
# ===========================================================================


# ===========================================================================
# HELPERS MATPLOTLIB — Formatage et style commun
# ===========================================================================

def _eur_axis_fmt(x: float, _: Any) -> str:
    """Formatte une valeur monétaire pour l'affichage sur un axe matplotlib.

    Utilise les suffixes k et M pour les grands montants afin d'éviter
    l'encombrement de l'axe avec des chiffres complets.

    Args:
        x: Valeur numérique.
        _: Position (ignorée, requise par l'API FuncFormatter de matplotlib).

    Returns:
        Chaîne formatée, ex: ``"1 250k €"``, ``"2.3M €"``, ``"850 €"``.
    """
    if abs(x) >= 1_000_000:
        return f"{x / 1_000_000:.1f}M €"
    if abs(x) >= 1_000:
        return f"{x / 1_000:.0f}k €"
    return f"{x:.0f} €"


def _pct_axis_fmt(x: float, _: Any) -> str:
    """Formatte une valeur en pourcentage pour l'affichage sur un axe matplotlib."""
    return f"{x:+.1f} %" if x != 0 else "0 %"


def _apply_chart_style(
    ax: Any,
    title: str,
    xlabel: str,
    ylabel: str,
    grid_axis: str = "y",
) -> None:
    """Applique la charte graphique sobre (navy/gris/blanc) à un axe matplotlib.

    Le titre est placé AU-DESSUS de la zone de tracé (comportement natif matplotlib),
    les labels d'axes sont À L'EXTÉRIEUR par défaut — ce qui résout le problème
    de chevauchement des graphiques openpyxl natifs.

    Args:
        ax: Axe matplotlib à styliser.
        title: Titre du graphique (au-dessus, aligné à gauche, navy gras).
        xlabel: Label de l'axe X (en dessous, gris).
        ylabel: Label de l'axe Y (à gauche, gris).
        grid_axis: Axe sur lequel afficher la grille (``"x"``, ``"y"`` ou ``"both"``).
    """
    ax.set_title(
        title,
        fontsize=10, fontweight="bold",
        color=f"#{_NAVY}", loc="left", pad=10,
    )
    ax.set_xlabel(xlabel, fontsize=8, color=f"#{_GREY_MID}", labelpad=6)
    ax.set_ylabel(ylabel, fontsize=8, color=f"#{_GREY_MID}", labelpad=6)
    ax.tick_params(
        axis="both", labelsize=7.5,
        labelcolor=f"#{_GREY_DARK}", length=3, width=0.5,
    )
    ax.grid(
        axis=grid_axis,
        color=f"#{_BORDER_COL}", linestyle="-", linewidth=0.5, alpha=0.8,
    )
    ax.set_axisbelow(True)
    ax.set_facecolor(f"#{_GREY_LITE}")
    for spine in ("top", "right"):
        ax.spines[spine].set_visible(False)
    for spine in ("left", "bottom"):
        ax.spines[spine].set_color(f"#{_BORDER_COL}")
        ax.spines[spine].set_linewidth(0.5)


def _fig_to_xl_image(fig: Any) -> XLImage:
    """Convertit une figure matplotlib en objet Image openpyxl (PNG en mémoire).

    La figure est fermée après conversion pour libérer la mémoire (important en
    mode Streamlit où plusieurs rapports peuvent être générés en séquence).
    La taille d'affichage dans Excel est alignée sur ``_CHART_W`` / ``_CHART_H``.

    Args:
        fig: Figure matplotlib à convertir.

    Returns:
        Objet ``XLImage`` prêt à être inséré avec ``ws.add_image()``.
    """
    buf = io.BytesIO()
    fig.savefig(
        buf, format="png", dpi=150,
        bbox_inches="tight", facecolor=fig.get_facecolor(),
    )
    plt.close(fig)
    buf.seek(0)
    img        = XLImage(buf)
    img.width  = int(_CHART_W / 2.54 * 96)   # cm → pixels (96 dpi référence Excel)
    img.height = int(_CHART_H / 2.54 * 96)
    return img


def _chart_ca_mensuel(months: pd.DataFrame) -> XLImage:
    """Graphique 1 — CA mensuel (barres verticales, bleu marine).

    Le graphique le plus fondamental pour un DAF. Permet de lire d'un coup
    d'œil les mois forts et les creux d'activité, base de toute discussion
    budgétaire ou de prévision.

    Args:
        months: DataFrame mensuel trié par mois [mois, montant].

    Returns:
        Image PNG openpyxl prête à être insérée dans la feuille Excel.
    """
    labels = months["mois"].astype(str).tolist()
    values = months["montant"].tolist()
    x      = range(len(labels))

    fig, ax = plt.subplots(figsize=(_CHART_W / 2.54, _CHART_H / 2.54))
    fig.patch.set_facecolor(f"#{_WHITE}")

    ax.bar(x, values, color=f"#{_NAVY}", width=0.6, zorder=3)
    ax.set_xticks(list(x))
    ax.set_xticklabels(labels, rotation=45, ha="right")
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(_eur_axis_fmt))

    _apply_chart_style(ax, "Chiffre d'affaires mensuel", "Mois", "Montant (€)")
    fig.tight_layout()
    return _fig_to_xl_image(fig)


def _chart_ca_cumule(months: pd.DataFrame) -> XLImage:
    """Graphique 2 — CA cumulé sur la période (courbe avec remplissage, bleu moyen).

    La courbe cumulée révèle la trajectoire de l'exercice : accélération,
    stagnation ou ralentissement. Indispensable pour les projections de
    fin d'année et la comparaison N vs N-1 (si plusieurs exercices).

    Args:
        months: DataFrame mensuel trié par mois [mois, montant].

    Returns:
        Image PNG openpyxl.
    """
    labels = months["mois"].astype(str).tolist()
    cumsum: list[float] = []
    s = 0.0
    for v in months["montant"].tolist():
        s += v
        cumsum.append(s)
    x = range(len(labels))

    fig, ax = plt.subplots(figsize=(_CHART_W / 2.54, _CHART_H / 2.54))
    fig.patch.set_facecolor(f"#{_WHITE}")

    ax.plot(list(x), cumsum, color=f"#{_NAVY_MID}", linewidth=2, zorder=3)
    ax.scatter(list(x), cumsum, color=f"#{_NAVY_MID}", s=25, zorder=4)
    ax.fill_between(list(x), cumsum, alpha=0.12, color=f"#{_NAVY_MID}")
    ax.set_xticks(list(x))
    ax.set_xticklabels(labels, rotation=45, ha="right")
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(_eur_axis_fmt))

    _apply_chart_style(ax, "CA cumulé sur la période", "Mois", "Montant cumulé (€)")
    fig.tight_layout()
    return _fig_to_xl_image(fig)


def _chart_variation_mom(months: pd.DataFrame) -> XLImage:
    """Graphique 3 — Variation mensuelle M/M en % (barres, vert/rouge selon signe).

    Signal d'alerte immédiat : détecte les ruptures de tendance mois par mois.
    Les barres sont colorées en vert (hausse) ou rouge (baisse) pour une lecture
    instantanée — avantage clé de matplotlib vs openpyxl natif qui ne supporte
    pas les couleurs conditionnelles par valeur.
    Le premier mois est exclu (pas de mois précédent disponible).

    Args:
        months: DataFrame mensuel trié par mois [mois, montant].

    Returns:
        Image PNG openpyxl.
    """
    raw_values = months["montant"].tolist()
    raw_labels = months["mois"].astype(str).tolist()

    # Calcul des variations — le premier mois n'a pas de précédent
    variations: list[float] = []
    v_labels:   list[str]   = []
    for i in range(1, len(raw_values)):
        if raw_values[i - 1] != 0:
            variations.append((raw_values[i] - raw_values[i - 1]) / abs(raw_values[i - 1]) * 100.0)
            v_labels.append(raw_labels[i])

    x      = range(len(v_labels))
    colors = [f"#{_GREEN}" if v >= 0 else f"#{_RED}" for v in variations]

    fig, ax = plt.subplots(figsize=(_CHART_W / 2.54, _CHART_H / 2.54))
    fig.patch.set_facecolor(f"#{_WHITE}")

    ax.bar(list(x), variations, color=colors, width=0.6, zorder=3)
    ax.axhline(y=0, color=f"#{_GREY_MID}", linewidth=0.8, zorder=2)
    ax.set_xticks(list(x))
    ax.set_xticklabels(v_labels, rotation=45, ha="right")
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(_pct_axis_fmt))

    _apply_chart_style(ax, "Variation mensuelle M/M (%)", "Mois", "Variation (%)")
    fig.tight_layout()
    return _fig_to_xl_image(fig)


def _chart_top_commerciaux(salespeople: pd.DataFrame) -> Optional[XLImage]:
    """Graphique 4 — Classement des commerciaux par CA (barres horizontales).

    Barres horizontales pour afficher les noms complets sans troncature.
    Les noms sont inversés pour que le meilleur apparaisse en haut.
    Limité aux 10 premiers pour la lisibilité.

    Args:
        salespeople: DataFrame [commercial, montant] trié par CA décroissant,
            ou DataFrame vide si pas de données commerciales.

    Returns:
        Image PNG openpyxl, ou None si ``salespeople`` est vide.
    """
    if salespeople.empty:
        return None

    top    = salespeople.head(10).reset_index(drop=True)
    # Inversion pour que le top 1 soit en haut du graphique
    labels = top["commercial"].astype(str).tolist()[::-1]
    values = top["montant"].tolist()[::-1]
    y      = range(len(labels))

    fig, ax = plt.subplots(figsize=(_CHART_W / 2.54, _CHART_H / 2.54))
    fig.patch.set_facecolor(f"#{_WHITE}")

    ax.barh(list(y), values, color=f"#{_NAVY_MID}", height=0.6, zorder=3)
    ax.set_yticks(list(y))
    ax.set_yticklabels(labels)
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(_eur_axis_fmt))

    _apply_chart_style(
        ax, "Classement des commerciaux par CA",
        "Montant (€)", "Commercial", grid_axis="x",
    )
    fig.tight_layout()
    return _fig_to_xl_image(fig)


def _chart_panier_moyen(months: pd.DataFrame, df: pd.DataFrame) -> XLImage:
    """Graphique 5 — Évolution du panier moyen mensuel (courbe marqueurs diamant).

    Indicateur qualité stratégiquement clé : un CA stable avec un panier en hausse
    signifie moins de clients mais mieux valorisés → signal de montée en gamme.
    Inversement, un panier en baisse avec CA stable signale une dilution du portefeuille.

    Args:
        months: DataFrame mensuel trié par mois [mois, montant].
        df: DataFrame nettoyé complet, pour calculer le nb de transactions par mois.

    Returns:
        Image PNG openpyxl.
    """
    tx_by_month = (
        df[df["montant"].notna()]
        .groupby("mois").size()
        .rename("nb_tx")
        .reset_index()
    )
    data          = months.merge(tx_by_month, on="mois", how="left")
    data["nb_tx"] = data["nb_tx"].fillna(1).clip(lower=1)
    data["panier"] = data["montant"] / data["nb_tx"]

    labels = data["mois"].astype(str).tolist()
    values = data["panier"].tolist()
    x      = range(len(labels))

    fig, ax = plt.subplots(figsize=(_CHART_W / 2.54, _CHART_H / 2.54))
    fig.patch.set_facecolor(f"#{_WHITE}")

    ax.plot(list(x), values, color=f"#{_GREY_DARK}", linewidth=2, zorder=3)
    ax.scatter(list(x), values, color=f"#{_GREY_DARK}", s=30, marker="D", zorder=4)
    ax.fill_between(list(x), values, alpha=0.08, color=f"#{_GREY_DARK}")
    ax.set_xticks(list(x))
    ax.set_xticklabels(labels, rotation=45, ha="right")
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(_eur_axis_fmt))

    _apply_chart_style(ax, "Panier moyen mensuel (€ / transaction)", "Mois", "Panier moyen (€)")
    fig.tight_layout()
    return _fig_to_xl_image(fig)


# ===========================================================================
# FEUILLE DASHBOARD — Mise en page KPIs + graphiques
# ===========================================================================

def _kpi_card(
    ws: Any,
    row: int,
    col: int,
    label: str,
    value: Any,
    context: str = "",
    value_color: str = _NAVY,
) -> None:
    """Écrit une carte KPI sur 3 lignes : libellé / valeur / contexte.

    La carte utilise des bordures asymétriques (haut seul, côtés seuls, bas seul)
    pour créer un contour propre autour des 3 lignes sans double trait entre elles.

    Args:
        ws: Feuille cible.
        row: Ligne du libellé (les lignes row+1 et row+2 sont occupées).
        col: Colonne de la carte.
        label: Libellé en petit gris en haut.
        value: Valeur principale en grand gras navy (ou couleur custom).
        context: Ligne de contexte en petit italique gris en bas.
        value_color: Couleur de la valeur (pour variations pos/neg).
    """
    thin = _THIN_SIDE

    _w(ws, row,   col, label,   size=9,  color=_GREY_MID,  align="center", bg=_GREY_LITE)
    _w(ws, row+1, col, value,   size=14, color=value_color, bold=True, align="center", bg=_GREY_LITE)
    _w(ws, row+2, col, context, size=8,  color=_GREY_MID,  align="center", bg=_GREY_LITE, italic=True)

    ws.cell(row,   col).border = Border(top=thin,    left=thin, right=thin)
    ws.cell(row+1, col).border = Border(left=thin,   right=thin)
    ws.cell(row+2, col).border = Border(bottom=thin, left=thin, right=thin)

    ws.row_dimensions[row].height   = 13
    ws.row_dimensions[row+1].height = 26
    ws.row_dimensions[row+2].height = 13


def _section_header(ws: Any, row: int, text: str, max_col: int = 6) -> None:
    """Écrit un en-tête de section navy fusionné sur plusieurs colonnes.

    Args:
        ws: Feuille cible.
        row: Numéro de ligne.
        text: Texte de l'en-tête.
        max_col: Dernière colonne de la fusion.
    """
    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row,   end_column=max_col,
    )
    _w(ws, row, 1, text, bold=True, size=11, color=_NAVY, align="left")
    ws.row_dimensions[row].height = 18


def _build_dashboard_sheet(
    wb: Any,
    report_months: pd.DataFrame,
    report_salespeople: pd.DataFrame,
    df: pd.DataFrame,
    kpis: DashboardKPIs,
) -> None:
    """Construit la feuille '📊 Dashboard' dans le classeur.

    Layout :
        Ligne 1    : titre principal (fond navy, texte blanc)
        Ligne 2    : sous-titre période + date de génération
        Ligne 3    : vide (respiration)
        Ligne 4    : en-tête section "INDICATEURS CLÉS"
        Lignes 5-7 : 3 cartes KPI (ligne A)
        Ligne 8    : vide
        Lignes 9-11: 3 cartes KPI (ligne B)
        Ligne 12   : vide
        Ligne 13   : en-tête section "ANALYSES GRAPHIQUES"
        Lignes 14+ : 5 graphiques PNG matplotlib (2 par ligne, 3 lignes)

    Args:
        wb: Classeur openpyxl.
        report_months: Agrégation mensuelle [mois, montant] pour la période.
        report_salespeople: Agrégation par commercial [commercial, montant] ou vide.
        df: DataFrame nettoyé complet (pour le calcul du panier moyen).
        kpis: Indicateurs pré-calculés par ``compute_dashboard_kpis``.
    """
    ws = wb.create_sheet("📊 Dashboard")

    # Largeurs de colonnes
    for col_idx, width in _COL_WIDTHS_DB.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    period_label = (
        f"{report_months['mois'].min()} → {report_months['mois'].max()}"
    )

    # -----------------------------------------------------------------------
    # Ligne 1 : Titre principal
    # -----------------------------------------------------------------------
    ws.merge_cells("A1:F1")
    _w(ws, 1, 1,
       f"TABLEAU DE BORD  ·  {SETTINGS.app_name}",
       bold=True, size=18, color=_WHITE, bg=_NAVY, align="center")
    ws.row_dimensions[1].height = 34

    # -----------------------------------------------------------------------
    # Ligne 2 : Sous-titre
    # -----------------------------------------------------------------------
    ws.merge_cells("A2:F2")
    _w(ws, 2, 1,
       f"Période : {period_label}   |   "
       f"Généré le {datetime.today().strftime('%d/%m/%Y')}   |   "
       f"v{SETTINGS.app_version}",
       size=9, color=_GREY_MID, bg=_GREY_LITE, align="center", italic=True)
    ws.row_dimensions[2].height = 15

    # Ligne 3 : respiration
    ws.row_dimensions[3].height = 8

    # -----------------------------------------------------------------------
    # Ligne 4 : En-tête section KPIs
    # -----------------------------------------------------------------------
    _section_header(ws, 4, "INDICATEURS CLÉS")

    # -----------------------------------------------------------------------
    # Lignes 5-7 : KPI Row A (CA total, Panier moyen, Nb transactions)
    # -----------------------------------------------------------------------
    _kpi_card(
        ws, 5, 1,
        "CHIFFRE D'AFFAIRES TOTAL",
        _fmt_eur(kpis.total_revenue),
        f"{kpis.active_months} mois couverts",
    )
    _kpi_card(
        ws, 5, 3,
        "PANIER MOYEN",
        _fmt_eur(kpis.average_basket, decimals=2),
        "par transaction valide",
    )
    _kpi_card(
        ws, 5, 5,
        "NB TRANSACTIONS VALIDES",
        f"{kpis.transaction_count:,}".replace(",", " "),
    )

    # Ligne 8 : respiration
    ws.row_dimensions[8].height = 8

    # -----------------------------------------------------------------------
    # Lignes 9-11 : KPI Row B (Variation M/M, Meilleur mois, Top commercial)
    # -----------------------------------------------------------------------

    # Variation M/M — couleur conditionnelle : vert si ≥ 0, rouge sinon
    if kpis.last_mom_growth_pct is not None:
        mom_val   = f"{kpis.last_mom_growth_pct:+.1f} %"
        mom_color = _GREEN if kpis.last_mom_growth_pct >= 0 else _RED
        mom_ctx   = "dernier mois vs précédent"
    else:
        mom_val   = "— N/A"
        mom_color = _GREY_MID
        mom_ctx   = "données insuffisantes (< 2 mois)"

    _kpi_card(ws, 9, 1, "VARIATION M/M", mom_val, mom_ctx, value_color=mom_color)

    _kpi_card(
        ws, 9, 3,
        "MEILLEUR MOIS",
        kpis.best_month,
        f"{_fmt_eur(kpis.best_month_revenue)}  ·  pire : {kpis.worst_month}",
    )

    if kpis.top_salesperson:
        top_ctx = (
            f"{_fmt_eur(kpis.top_salesperson_revenue or 0)}  ·  "
            f"{kpis.active_salespeople} commerciaux"
        )
        _kpi_card(ws, 9, 5, "TOP COMMERCIAL", kpis.top_salesperson, top_ctx)
    else:
        _kpi_card(ws, 9, 5, "TOP COMMERCIAL", "—", "aucune donnée commerciale")

    # Ligne 12 : respiration
    ws.row_dimensions[12].height = 10

    # -----------------------------------------------------------------------
    # Ligne 13 : En-tête section graphiques
    # -----------------------------------------------------------------------
    _section_header(ws, 13, "ANALYSES GRAPHIQUES")

    # -----------------------------------------------------------------------
    # Graphiques : 2 par ligne (ancres G = col 7 pour les graphiques de droite)
    # Chaque graphique occupe ~18 lignes à hauteur standard.
    # -----------------------------------------------------------------------
    months_sorted = report_months.sort_values("mois").reset_index(drop=True)

    img1 = _chart_ca_mensuel(months_sorted)
    img2 = _chart_ca_cumule(months_sorted)
    img3 = _chart_variation_mom(months_sorted)
    img4 = _chart_top_commerciaux(report_salespeople)
    img5 = _chart_panier_moyen(months_sorted, df)

    ws.add_image(img1, "A14")
    ws.add_image(img2, "H14")

    ws.add_image(img3, "A35")
    if img4 is not None:
        ws.add_image(img4, "H35")
    else:
        # Texte d'absence si pas de colonne commercial dans les sources
        ws.merge_cells("G33:K40")
        _w(ws, 33, 7,
           "Graphique 4 non disponible :\naucune colonne « commercial » dans les fichiers sources.",
           size=10, color=_GREY_MID, align="center", italic=True, bg=_GREY_LITE, wrap=True)

    ws.add_image(img5, "A56")


# ===========================================================================
# FEUILLE RÉCONCILIATION — Rapport d'intégrité formaté
# ===========================================================================

def _build_reconciliation_sheet(
    wb: Any,
    report: ReconciliationReport,
) -> None:
    """Construit la feuille '🔐 Réconciliation' avec le rapport d'intégrité.

    Cette feuille est le miroir Excel du fichier log.txt : elle rend le
    checkpoint d'intégrité lisible directement dans le classeur, sans
    supprimer le log (les deux coexistent pour la traçabilité complète).

    Args:
        wb: Classeur openpyxl.
        report: Rapport de réconciliation issu de ``utils.reconcile_data()``.
    """
    ws = wb.create_sheet("🔐 Réconciliation")

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 5   # col de décoration

    # --- Titre ---
    ws.merge_cells("A1:B1")
    _w(ws, 1, 1, "RAPPORT DE RÉCONCILIATION DES DONNÉES",
       bold=True, size=14, color=_WHITE, bg=_NAVY, align="center")
    ws.row_dimensions[1].height = 30

    # --- Sous-titre ---
    ws.merge_cells("A2:B2")
    _w(ws, 2, 1,
       f"Généré le {datetime.today().strftime('%d/%m/%Y à %H:%M')}  |  "
       f"{SETTINGS.app_name} v{SETTINGS.app_version}",
       size=9, color=_GREY_MID, bg=_GREY_LITE, align="center", italic=True)
    ws.row_dimensions[2].height = 14

    ws.row_dimensions[3].height = 8  # respiration

    # --- Résultat global ---
    _section_header(ws, 4, "RÉSULTAT DU CONTRÔLE", max_col=2)

    status_text = "✅  INTÉGRITÉ VALIDÉE" if report.integrity_ok else "🚨  ALERTE — ÉCART DÉTECTÉ"
    status_bg   = _GREY_LITE if report.integrity_ok else _ALERT_BG
    status_col  = _GREEN     if report.integrity_ok else _RED

    ws.merge_cells("A5:B5")
    _w(ws, 5, 1, status_text,
       bold=True, size=13, color=status_col, bg=status_bg, align="center")
    ws.cell(5, 1).border = _THIN_BORDER
    ws.row_dimensions[5].height = 28

    ws.merge_cells("A6:B6")
    _w(ws, 6, 1, report.message,
       size=9, color=_GREY_DARK, bg=status_bg, align="left", wrap=True)
    ws.row_dimensions[6].height = 32

    ws.row_dimensions[7].height = 8

    # --- Tableau des montants ---
    _section_header(ws, 8, "DÉTAIL DES MONTANTS", max_col=2)

    money_fmt = '#,##0.00 "€"'
    amounts_rows = [
        ("Somme source attendue (lignes à date valide)", report.expected_source_sum, money_fmt),
        ("Somme traitée après nettoyage",                report.processed_sum,       money_fmt),
        ("Écart absolu",                                 report.absolute_gap,        money_fmt),
        ("Écart relatif (%)",                            report.gap_pct * 100,       '0.0000"%"'),
        ("Tolérance appliquée (%)",
         SETTINGS.reconciliation_tolerance_pct * 100,   '0.00"%"'),
    ]

    for i, (label, value, fmt) in enumerate(amounts_rows):
        r      = 9 + i
        row_bg = _GREY_LITE if i % 2 == 0 else _WHITE
        _w(ws, r, 1, label, size=10, color=_GREY_DARK, bg=row_bg)
        ws.cell(r, 1).border = _THIN_BORDER
        cell = ws.cell(r, 2)
        cell.value         = value
        cell.font          = _font(bold=True, size=10, color=_NAVY)
        cell.fill          = _fill(row_bg)
        cell.alignment     = Alignment(horizontal="right", vertical="center")
        cell.border        = _THIN_BORDER
        cell.number_format = fmt
        ws.row_dimensions[r].height = 18

    ws.row_dimensions[14].height = 8

    # --- Tableau du comptage des lignes ---
    _section_header(ws, 15, "DÉTAIL DES LIGNES", max_col=2)

    lines_rows = [
        ("Lignes sources (avant nettoyage)",           report.source_row_count,    False),
        ("Lignes après nettoyage",                     report.processed_row_count, False),
        ("Lignes supprimées (total)",                  report.dropped_row_count,   False),
        ("  → dont dates invalides (parse impossible)", report.invalid_date_count,  True),
        ("  → dont montants non convertibles (→ NaN)", report.invalid_amount_count, True),
    ]

    for i, (label, value, is_sub) in enumerate(lines_rows):
        r      = 16 + i
        row_bg = _GREY_LITE if i % 2 == 0 else _WHITE
        _w(ws, r, 1, label, size=10,
           color=_GREY_MID if is_sub else _GREY_DARK,
           italic=is_sub, bg=row_bg)
        ws.cell(r, 1).border = _THIN_BORDER
        _w(ws, r, 2, value, bold=True, size=10, color=_NAVY, align="right", bg=row_bg)
        ws.cell(r, 2).border = _THIN_BORDER
        ws.row_dimensions[r].height = 16

    ws.row_dimensions[21].height = 8

    # --- Note d'audit ---
    ws.merge_cells("A22:B22")
    _w(ws, 22, 1,
       "ℹ️  Ce rapport est également disponible dans le fichier log.txt généré lors de "
       "l'exécution. Les deux documents sont complémentaires pour l'audit.",
       size=8, color=_GREY_MID, italic=True, align="left", wrap=True)
    ws.row_dimensions[22].height = 28


# ===========================================================================
# FORMATAGE DES FEUILLES DE DONNÉES
# ===========================================================================

def _format_data_sheet(ws: Any) -> None:
    """Applique le formatage standard aux feuilles de données du classeur enrichi.

    Variante de ``utils._format_excel_sheet`` avec en-têtes navy et alternance
    de fond sur les lignes pour cohérence avec la charte graphique du dashboard.
    Volontairement indépendante de la fonction privée de utils.py pour respecter
    le principe d'encapsulation (dashboard.py est un module autonome).

    Args:
        ws: Feuille openpyxl à formater (modifiée en place).
    """
    from openpyxl.styles import numbers as openpyxl_numbers

    if ws.max_row < 1:
        return

    # En-têtes : fond bleu marine, texte blanc, gras, centré
    for cell in ws[1]:
        cell.font      = _font(bold=True, size=10, color=_WHITE)
        cell.fill      = _fill(_NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Ligne 1 figée
    ws.freeze_panes = "A2"

    # Largeurs automatiques avec planchers métier
    headers   = [str(c.value).lower() if c.value else "" for c in ws[1]]
    for col_cells in ws.columns:
        col_letter = col_cells[0].column_letter
        header_val = str(col_cells[0].value) if col_cells[0].value else ""
        max_len    = max(
            len(header_val),
            max(
                (len(str(c.value)) for c in col_cells[1:] if c.value is not None),
                default=0,
            ),
        )
        width = min(max(max_len + 4, 12), 40)
        if any(kw in header_val.lower() for kw in ("montant", "euro", "total", "ca")):
            width = max(width, 18)
        ws.column_dimensions[col_letter].width = width

    # Alternance de fond gris très clair / blanc sur les lignes de données
    for idx, row_cells in enumerate(ws.iter_rows(min_row=2), start=2):
        row_bg = _GREY_LITE if idx % 2 == 0 else _WHITE
        for cell in row_cells:
            cell.fill = _fill(row_bg)

    # Formats date et monétaire
    header_idx = {h: i + 1 for i, h in enumerate(headers) if h}
    if "date" in header_idx:
        c = header_idx["date"]
        for r in range(2, ws.max_row + 1):
            ws.cell(r, c).number_format = "DD/MM/YYYY"
    for kw in ("montant en euros", "montant", "total", "ca"):
        if kw in header_idx:
            c = header_idx[kw]
            for r in range(2, ws.max_row + 1):
                ws.cell(r, c).number_format = openpyxl_numbers.FORMAT_CURRENCY_EUR_SIMPLE
            break


# ===========================================================================
# ASSEMBLAGE DU CLASSEUR COMPLET
# ===========================================================================

def _assemble_workbook(
    df: pd.DataFrame,
    report_months: pd.DataFrame,
    report_salespeople: pd.DataFrame,
    reconciliation_report: ReconciliationReport,
) -> Any:
    """Assemble le classeur Excel enrichi avec toutes les feuilles dans l'ordre.

    Ordre final des feuilles (visible pour l'utilisateur) :
        1. 📊 Dashboard
        2. 📅 Par mois
        3. 👤 Par commercial
        4. 🔍 Données brutes
        5. 🔐 Réconciliation
        (_data est cachée et reste en dernière position)

    Args:
        df: DataFrame nettoyé.
        report_months: Agrégation mensuelle [mois, montant].
        report_salespeople: Agrégation par commercial [commercial, montant].
        reconciliation_report: Rapport de réconciliation issu de ``utils.reconcile_data()``.

    Returns:
        Classeur openpyxl complet, ordonné et formaté, prêt à être sauvegardé.
    """
    # Copies pour l'export : renommage de la colonne interne "montant"
    # en "montant en euros" pour la lisibilité dans les feuilles de données.
    months_xp     = report_months.copy().rename(columns={"montant": "montant en euros"})
    salespeople_xp = report_salespeople.copy().rename(columns={"montant": "montant en euros"})

    # Création initiale via pandas ExcelWriter (gère la sérialisation des DataFrames)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        months_xp.to_excel(     writer, sheet_name="📅 Par mois",        index=False)
        salespeople_xp.to_excel(writer, sheet_name="👤 Par commercial",  index=False)
        df.to_excel(            writer, sheet_name="🔍 Données brutes",   index=False)
    buf.seek(0)
    wb = load_workbook(buf)

    # Formatage des feuilles de données
    for name in ["📅 Par mois", "👤 Par commercial", "🔍 Données brutes"]:
        if name in wb.sheetnames:
            _format_data_sheet(wb[name])

    # Calcul des KPIs
    kpis = compute_dashboard_kpis(df, report_months, report_salespeople)

    # Ajout des nouvelles feuilles
    _build_dashboard_sheet(wb, report_months, report_salespeople, df, kpis)
    _build_reconciliation_sheet(wb, reconciliation_report)

    # Réordonnancement : Dashboard en premier, _data cachée en dernier.
    # Manipulation de wb._sheets (liste interne openpyxl 3.x) :
    # l'API publique move_sheet() ne permet pas de spécifier un ordre absolu.
    target_order = [
        "📊 Dashboard",
        "📅 Par mois",
        "👤 Par commercial",
        "🔍 Données brutes",
        "🔐 Réconciliation",
    ]
    sheet_map  = {ws.title: ws for ws in wb.worksheets}
    wb._sheets = [sheet_map[name] for name in target_order if name in sheet_map]

    return wb


# ===========================================================================
# FONCTIONS PUBLIQUES — API Streamlit et Batch
# ===========================================================================

def build_excel_bytes_with_dashboard(
    df: pd.DataFrame,
    report_months: pd.DataFrame,
    report_salespeople: pd.DataFrame,
    reconciliation_report: ReconciliationReport,
) -> bytes:
    """Génère le classeur Excel enrichi en mémoire (mode Streamlit).

    Remplace le couple ``build_excel_bytes`` + ``format_excel_bytes`` de utils.py.
    Toutes les feuilles sont créées, formatées et ordonnées en un seul passage.

    Args:
        df: DataFrame nettoyé (colonnes : date, montant, mois, commercial optionnel).
        report_months: Agrégation mensuelle [mois, montant].
        report_salespeople: Agrégation par commercial [commercial, montant] ou vide.
        reconciliation_report: Rapport issu de ``utils.reconcile_data()``.

    Returns:
        Bytes du fichier ``.xlsx`` complet, prêt à être proposé en téléchargement.
    """
    wb  = _assemble_workbook(df, report_months, report_salespeople, reconciliation_report)
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    logging.info(
        "Classeur Excel enrichi généré en mémoire "
        "(dashboard + réconciliation + données formatées)."
    )
    return out.getvalue()


def export_excel_with_dashboard(
    df: pd.DataFrame,
    report_months: pd.DataFrame,
    report_salespeople: pd.DataFrame,
    reconciliation_report: ReconciliationReport,
    filepath: Path,
) -> Path:
    """Exporte le classeur Excel enrichi sur le disque (mode Batch).

    Remplace le couple ``export_excel`` + ``format_excel_file`` de utils.py.
    Crée le répertoire parent si nécessaire.

    Args:
        df: DataFrame nettoyé.
        report_months: Agrégation mensuelle.
        report_salespeople: Agrégation par commercial.
        reconciliation_report: Rapport de réconciliation.
        filepath: Chemin de destination (fichier ``.xlsx``).

    Returns:
        Path du fichier créé.

    Raises:
        Exception: Propagée si l'écriture sur disque échoue (disque plein, droits…).
    """
    Path(filepath).parent.mkdir(parents=True, exist_ok=True)
    try:
        wb = _assemble_workbook(
            df, report_months, report_salespeople, reconciliation_report
        )
        wb.save(str(filepath))
        logging.info(f"Classeur Excel enrichi exporté : {filepath}")
        return Path(filepath)
    except Exception as exc:
        logging.error(f"Erreur export classeur enrichi : {filepath} | {exc}")
        raise