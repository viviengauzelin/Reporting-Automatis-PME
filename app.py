from datetime import datetime
import logging
import hashlib
import io

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, numbers

from utils import (
    lire_excels_upload,  # DOIT retourner (df, fichiers_ko)
    nettoyer,
    ajouter_mois,
    total_par_mois,
    total_par_commercial,
    build_excel_bytes,
    build_pdf_bytes,
)

# Réduire le bruit des logs terminal (Streamlit relance souvent le script)
logging.getLogger().setLevel(logging.ERROR)

st.set_page_config(page_title="Reporting (Excel)", layout="centered")
st.title("📊 Reporting — Excel-only")
st.write("Upload des .xlsx → mapping colonnes → génération → téléchargement Excel + PDF + log.")


# -----------------------------
# Helpers
# -----------------------------
def normalize_and_merge_duplicate_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Normalise les noms de colonnes (strip + lower) puis fusionne les colonnes dupliquées (bfill)."""
    df = df.copy()
    df.columns = df.columns.astype(str).str.strip().str.lower()

    if df.columns.duplicated().any():
        dup_names = df.columns[df.columns.duplicated()].unique().tolist()
        for name in dup_names:
            cols = df.loc[:, df.columns == name]
            merged = cols.bfill(axis=1).iloc[:, 0]  # prend la 1ère valeur non vide parmi les doublons
            df = df.drop(columns=name)
            df[name] = merged
    return df


def default_index(cols: list[str], preferred: list[str], fallback: int = 0) -> int:
    # Pré-sélection “intelligente” dans le selectbox si une colonne attendue existe
    for p in preferred:
        if p in cols:
            return cols.index(p)
    return fallback


def fmt_eur_fr(x: float) -> str:
    # Transformer un nombre (float) en format monétaire français.
    # Exemple : 1234567.89 -> 1234567.89 -> 1 234 567,89 €
    s = f"{x:,.2f}"
    s = s.replace(",", "X").replace(".", ",").replace("X", " ")
    return f"{s} €"


def make_log_text(lines: list[str]) -> str:
    # Transformer une liste de lignes en un texte complet prêt à être écrit dans un fichier log.
    return "\n".join(lines) + "\n"


def format_excel_bytes_like_batch(excel_bytes: bytes) -> bytes:
    """
    Applique un formatage type 'batch' (headers gras, freeze, largeur, formats date/€)
    directement sur le fichier Excel en mémoire.
    """
    buf = io.BytesIO(excel_bytes)  # fichier “virtuel” en RAM
    wb = load_workbook(buf)

    for ws in wb.worksheets:
        # Header en gras
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Ligne 1 figée
        ws.freeze_panes = "A2"

        # Ajustement largeur colonnes (robuste)
        for col_cells in ws.columns:
            col_letter = col_cells[0].column_letter
            header = str(col_cells[0].value) if col_cells[0].value else ""

            max_len = len(header)
            for cell in col_cells[1:]:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))

            width = max_len + 4
            width = max(width, 12)

            if "montant" in header.lower() or "euro" in header.lower():
                width = max(width, 18)

            width = min(width, 40)
            ws.column_dimensions[col_letter].width = width

        # Formats date / monnaie
        headers = [cell.value for cell in ws[1]]
        header_to_index = {h: i + 1 for i, h in enumerate(headers) if h}

        if "date" in header_to_index:
            c = header_to_index["date"]
            for r in range(2, ws.max_row + 1):
                ws.cell(r, c).number_format = "DD/MM/YYYY"

        for money_col in ("montant", "montant en euros", "total"):
            if money_col in header_to_index:
                c = header_to_index[money_col]
                for r in range(2, ws.max_row + 1):
                    ws.cell(r, c).number_format = numbers.FORMAT_CURRENCY_EUR_SIMPLE

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def hash_uploaded_file(f) -> str:
    # Empreinte unique du contenu du fichier (audit / anti-litige)
    b = f.getvalue()
    return hashlib.sha256(b).hexdigest()


def compute_data_quality_metrics(df_mapped: pd.DataFrame) -> dict:
    """
    Métriques qualité AVANT nettoyage final :
    - Dates invalides (format incorrect ou date inexistante)
    - montants vides (source)
    - montants invalides hors vides (non convertibles alors qu'il y a une valeur)
    """
    metrics = {
        "dates_invalides": 0,
        "montants_vides_source": 0,
        "montants_invalides_hors_vides": 0,
    }

    if "date" in df_mapped.columns:
        date_parsed = pd.to_datetime(df_mapped["date"], dayfirst=True, errors="coerce")
        metrics["dates_invalides"] = int(date_parsed.isna().sum())

    if "montant" in df_mapped.columns:
        raw = df_mapped["montant"]
        raw_str = raw.astype(str).str.strip()
        mask_vide = raw.isna() | (raw_str == "")
        metrics["montants_vides_source"] = int(mask_vide.sum())

        s = raw_str.str.replace(",", ".", regex=False)
        s = s.str.replace("\u00a0", "", regex=False)
        num = pd.to_numeric(s, errors="coerce")

        mask_invalide_hors_vide = num.isna() & (~mask_vide)
        metrics["montants_invalides_hors_vides"] = int(mask_invalide_hors_vide.sum())

    return metrics


def friendly_error_message(err: str) -> str:
    """
    Transforme un message d'erreur technique en message compréhensible PME.
    """
    e = (err or "").lower()

    if "excel file format cannot be determined" in e:
        return "Format Excel non reconnu (fichier probablement cassé, mauvais format, ou renommé en .xlsx)."
    if "password" in e or "encrypted" in e:
        return "Fichier protégé par mot de passe / chiffré (impossible à lire automatiquement)."
    if "permission" in e or "access is denied" in e:
        return "Accès refusé (droits / fichier verrouillé)."
    if "zipfile" in e:
        return "Fichier Excel corrompu (structure interne illisible)."
    if "openpyxl" in e:
        return "Erreur de lecture Excel (fichier potentiellement incompatible / corrompu)."

    return "Le fichier n’est pas un vrai .xlsx (souvent : fichier renommé en .xlsx ou fichier corrompu)."


# -----------------------------
# Upload + déduplication
# -----------------------------
uploaded = st.file_uploader(
    "Fichiers Excel à consolider (.xlsx)",
    type=["xlsx"],
    accept_multiple_files=True,
)

if not uploaded:
    st.info("Ajoute 1+ fichiers Excel pour commencer.")
    st.stop()

# Dédupliquer les fichiers uploadés (par hash contenu)
unique_files = []
seen_hash = {}
duplicates = []

for f in uploaded:
    h = hash_uploaded_file(f)  # même fichier (même contenu) => même hash
    if h in seen_hash:
        duplicates.append(f.name)
    else:
        seen_hash[h] = f.name
        unique_files.append(f)

if duplicates:
    st.warning(f"Doublons ignorés (même contenu) : {', '.join(duplicates)}")

# Lecture brut (robuste) -> df_raw, fichiers_ko
df_raw, fichiers_ko = lire_excels_upload(unique_files)

# Affichage PME friendly des fichiers ignorés
if fichiers_ko:
    st.warning("⚠️ Certains fichiers ont été ignorés car ils n'ont pas pu être lus :")
    for name, err in fichiers_ko:
        st.write(f"- **{name}** — {friendly_error_message(err)}")

    with st.expander("Détails techniques (pour debug)", expanded=False):
        for name, err in fichiers_ko:
            st.write(f"- {name} — {err}")

nb_fichiers_ok = len(unique_files) - len(fichiers_ko)
nb_lignes_avant = len(df_raw)

st.success(f"{nb_fichiers_ok} fichier(s) OK lus. Total lignes (avant nettoyage) : {nb_lignes_avant}")

# Pré-nettoyage pour mapping: normaliser + fusionner colonnes dupliquées
df_base = normalize_and_merge_duplicate_columns(df_raw)
cols = list(df_base.columns)

# -----------------------------
# Form mapping stable
# -----------------------------
st.subheader("🧩 Mapping des colonnes")

with st.form("mapping_form"):
    col_date = st.selectbox("Colonne DATE", options=cols, index=default_index(cols, ["date"]))
    col_montant = st.selectbox("Colonne MONTANT", options=cols, index=default_index(cols, ["montant"]))
    col_commercial = st.selectbox(
        "Colonne COMMERCIAL (optionnel)",
        options=["(aucune)"] + cols,
        index=(["(aucune)"] + cols).index("commercial") if "commercial" in cols else 0,
    )
    submitted = st.form_submit_button("⚙️ Générer le reporting")  # form = évite rerun à chaque click/choix

# Bloque tant que l'utilisateur n'a pas lancé la génération au moins une fois.
if not submitted:
    if "generated" not in st.session_state:
        # "generated" est défini plus tard dans le code
        st.stop()

if submitted and col_date == col_montant:
    st.error("Tu as sélectionné la même colonne pour DATE et MONTANT. Choisis une colonne Montant.")
    st.stop()


# -----------------------------
# Génération + persistance de l'affichage
# -----------------------------
if submitted:
    rename_map = {col_date: "date", col_montant: "montant"}
    if col_commercial != "(aucune)":
        rename_map[col_commercial] = "commercial"

    df_mapped = df_base.rename(columns=rename_map)
    quality = compute_data_quality_metrics(df_mapped)

    # Log client (sera téléchargeable)
    log_lines = []
    now = datetime.now()

    log_lines.append("=== LOG D'EXECUTION (STREAMLIT) ===")
    log_lines.append(f"Horodatage: {now.strftime('%Y-%m-%d %H:%M:%S')}")
    log_lines.append(f"Fichiers uploadés (uniques): {len(unique_files)}")
    log_lines.append(f"Fichiers OK lus: {nb_fichiers_ok}")

    # ✅ Ajout audit: hash des fichiers dans le log
    log_lines.append("Liste fichiers uploadés (avec hash):")
    for f in unique_files:
        log_lines.append(f"  - {getattr(f, 'name', '(nom inconnu)')} | SHA256: {hash_uploaded_file(f)}")

    log_lines.append("")
    log_lines.append("=== FICHIERS EN ERREUR (ignores) ===")
    if fichiers_ko:
        for name, err in fichiers_ko:
            log_lines.append(f"- {name} | {err}")
    else:
        log_lines.append("Aucun")

    log_lines.append("")
    log_lines.append(f"Lignes (avant nettoyage): {nb_lignes_avant}")

    log_lines.append("")
    log_lines.append("=== MAPPING ===")
    log_lines.append(f"DATE: {col_date} -> date")
    log_lines.append(f"MONTANT: {col_montant} -> montant")
    log_lines.append(f"COMMERCIAL: {col_commercial} -> {'commercial' if col_commercial != '(aucune)' else '(aucune)'}")

    log_lines.append("")
    log_lines.append("=== QUALITE DES DONNEES (avant nettoyage final) ===")
    log_lines.append(f"Dates invalides (format incorrect ou date inexistante): {quality['dates_invalides']}")
    log_lines.append(f"Montants vides (source): {quality['montants_vides_source']}")
    log_lines.append(f"Montants invalides (hors vides): {quality['montants_invalides_hors_vides']}")

    try:
        df = nettoyer(df_mapped)
        df = ajouter_mois(df)

        report_mois = total_par_mois(df)
        report_commercial = total_par_commercial(df)

        nb_lignes_apres = len(df)
        nb_supprimees = nb_lignes_avant - nb_lignes_apres
        periode_min = df["mois"].min()
        periode_max = df["mois"].max()
        total_global = float(df["montant"].sum())
        nb_commerciaux = int(df["commercial"].nunique()) if "commercial" in df.columns else 0

        tag = f"{periode_min}" if periode_min == periode_max else f"{periode_min}_to_{periode_max}"

        excel_bytes = build_excel_bytes(df, report_mois, report_commercial)
        excel_bytes = format_excel_bytes_like_batch(excel_bytes)
        pdf_bytes = build_pdf_bytes(report_mois, report_commercial)

        log_lines.append("")
        log_lines.append("=== RESULTATS ===")
        log_lines.append(f"Lignes (après nettoyage): {nb_lignes_apres}")
        log_lines.append(f"Lignes supprimées: {nb_supprimees}")
        log_lines.append("Règle: lignes sans date valide supprimées ; montants vides conservés (NaN) et exclus du CA")
        log_lines.append(f"Période détectée: {periode_min} -> {periode_max}")
        log_lines.append(f"Chiffre d'affaires total: {fmt_eur_fr(total_global)}")
        if col_commercial != "(aucune)":
            log_lines.append(f"Nombre de commerciaux: {nb_commerciaux}")

        log_bytes = make_log_text(log_lines).encode("utf-8")

        st.session_state.generated = True

        # On stocke un statut pour éviter de crasher l'affichage si une exception arrive
        st.session_state.status = "ok"
        st.session_state.error_message = None

        st.session_state.tag = tag
        st.session_state.df = df
        st.session_state.report_mois = report_mois
        st.session_state.report_commercial = report_commercial
        st.session_state.summary = {
            "nb_fichiers_ok": nb_fichiers_ok,
            "nb_lignes_avant": nb_lignes_avant,
            "nb_lignes_apres": nb_lignes_apres,
            "nb_supprimees": nb_supprimees,
            "periode_min": periode_min,
            "periode_max": periode_max,
            "total_global": total_global,
            "nb_commerciaux": nb_commerciaux,
            "has_commercial": (col_commercial != "(aucune)"),
        }
        st.session_state.quality = quality
        st.session_state.excel_bytes = excel_bytes
        st.session_state.pdf_bytes = pdf_bytes
        st.session_state.log_bytes = log_bytes

    except Exception as e:
        st.error(f"Erreur: {e}")
        log_lines.append("")
        log_lines.append("=== ERREUR ===")
        log_lines.append(str(e))

        st.session_state.generated = True

        # Statut "error" => l'affichage en bas ne tente pas d'utiliser summary/df inexistants
        st.session_state.status = "error"
        st.session_state.error_message = str(e)

        st.session_state.tag = now.strftime("%Y-%m-%d_%H-%M-%S")
        st.session_state.log_bytes = make_log_text(log_lines).encode("utf-8")
        st.stop()


# -----------------------------
# Affichage persistant
# -----------------------------
if "generated" not in st.session_state:
    st.stop()

tag = st.session_state.tag

# ✅ Si erreur => on n'essaie pas d'afficher summary/df
if st.session_state.get("status") == "error":
    st.subheader("❌ Génération impossible")
    st.error(st.session_state.get("error_message", "Erreur inconnue."))

    st.download_button(
        "⬇️ Télécharger le log d'exécution",
        data=st.session_state.log_bytes,
        file_name=f"log_{tag}.txt",
        mime="text/plain",
    )
    st.stop()

s = st.session_state.summary
q = st.session_state.quality

st.subheader("✅ Résumé")
st.write(f"- Fichiers OK lus : **{s['nb_fichiers_ok']}**")
st.write(f"- Lignes avant nettoyage : **{s['nb_lignes_avant']}**")
st.write(f"- Lignes après nettoyage : **{s['nb_lignes_apres']}**")
st.write(f"- Lignes supprimées : **{s['nb_supprimees']}**")
st.write("- Règle : **lignes sans date valide supprimées** ; montants vides conservés (NaN) et **exclus du CA**.")
st.write(f"- Période détectée : **{s['periode_min']} → {s['periode_max']}**")
st.write(f"- Chiffre d’affaires total : **{fmt_eur_fr(s['total_global'])}**")
if s["has_commercial"]:
    st.write(f"- Nombre de commerciaux : **{s['nb_commerciaux']}**")

st.subheader("🧪 Qualité des données")
st.write(f"- Dates invalides (parse impossible) : **{q['dates_invalides']}**")
st.write(f"- Montants vides (source) : **{q['montants_vides_source']}**")
st.write(f"- Montants invalides (hors vides) : **{q['montants_invalides_hors_vides']}**")

st.subheader("Aperçu (après nettoyage)")
st.dataframe(st.session_state.df.head(30), width="stretch")
st.subheader("Total par mois")
st.dataframe(st.session_state.report_mois, width="stretch")

if s["has_commercial"]:
    st.subheader("Total par commercial")
    st.dataframe(st.session_state.report_commercial, width="stretch")

st.download_button(
    "⬇️ Télécharger l'Excel multi-feuilles (formaté)",
    data=st.session_state.excel_bytes,
    file_name=f"reporting_{tag}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

st.download_button(
    "⬇️ Télécharger le PDF",
    data=st.session_state.pdf_bytes,
    file_name=f"rapport_{tag}.pdf",
    mime="application/pdf",
)

st.download_button(
    "⬇️ Télécharger le log d'exécution",
    data=st.session_state.log_bytes,
    file_name=f"log_{tag}.txt",
    mime="text/plain",
)
