"""
utils.py - Moteur de traitement : chargement, nettoyage, audit, exports.

Ce module implémente la logique métier pure, indépendante de toute interface
(Streamlit ou Batch). Il est conçu pour être :

- **Auditable** : chaque transformation est tracée ; une fonction de réconciliation
  prouve mathématiquement que 100 % des données sources sont comptabilisées.
- **Robuste** : gestion explicite des erreurs, types incohérents, fichiers corrompus,
  encodages CSV variés (utf-8, latin-1, cp1252, utf-8-sig).
- **Testable** : aucune dépendance à Streamlit ou au système de fichiers dans le
  cœur métier (injection de dépendances).
- **Scalable** : voir note de scalabilité dans ``load_source_files``.

Dépendances :
    pandas, openpyxl, reportlab  (voir requirements.txt)
    config.SETTINGS, config.DATA_DICTIONARY  (ce projet)
"""

from __future__ import annotations

import hashlib
import io
import logging
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Optional, Protocol, runtime_checkable

import pandas as pd

from config import DATA_DICTIONARY, REQUIRED_COLUMNS, SETTINGS, ColSpec

# ---------------------------------------------------------------------------
# ALIAS DE TYPES
# ---------------------------------------------------------------------------

# Liste des fichiers ayant échoué : (nom_fichier, message_erreur)
FailedFiles = list[tuple[str, str]]

# Résultat du chargement Streamlit : (dataframe_consolidé, fichiers_en_erreur)
LoadResult = tuple[pd.DataFrame, FailedFiles]

# Extensions de fichiers sources reconnues (minuscules, sans point)
_SOURCE_EXTENSIONS: frozenset[str] = frozenset(SETTINGS.csv_source_extensions)


@runtime_checkable
class UploadedFileProtocol(Protocol):
    """Interface minimale attendue pour un fichier uploadé (ex: Streamlit UploadedFile).

    Définir un Protocol plutôt qu'importer UploadedFile de Streamlit permet de
    garder utils.py indépendant de Streamlit → testable avec de simples io.BytesIO.
    """

    name: str  # nom du fichier (ex: "ventes_2025-01.xlsx" ou "export_crm.csv")


# ---------------------------------------------------------------------------
# LOGGING
# ---------------------------------------------------------------------------

def setup_logging(base_dir: str = "output") -> Path:
    """Configure le logging Python vers fichier horodaté + console.

    Crée le répertoire ``base_dir`` s'il n'existe pas.
    ``force=True`` réinitialise tout handler existant (évite les doublons
    lors des relances en mode Batch/tests).

    Args:
        base_dir: Chemin du répertoire de sortie des logs.

    Returns:
        Path du fichier log créé (ex: ``output/log_2025-01-15.txt``).
    """
    Path(base_dir).mkdir(parents=True, exist_ok=True)
    today = datetime.today().strftime("%Y-%m-%d")
    log_path = Path(base_dir) / f"log_{today}.txt"

    logging.basicConfig(
        level=SETTINGS.log_level,
        format="%(asctime)s | %(levelname)s | %(message)s",
        handlers=[
            logging.FileHandler(log_path, encoding="utf-8"),
            logging.StreamHandler(),
        ],
        force=True,
    )
    return log_path


# ---------------------------------------------------------------------------
# HACHAGE - Intégrité des fichiers sources (audit)
# ---------------------------------------------------------------------------

def hash_file_on_disk(path: Path) -> str:
    """Calcule l'empreinte SHA-256 d'un fichier sur le disque.

    Le hash SHA-256 garantit qu'un fichier n'a pas été modifié entre son
    upload et son traitement. Toute altération (même un octet) produit un hash
    différent → détection de falsification ou de corruption.

    Lecture par blocs de 64 Ko pour éviter de charger tout le fichier en RAM
    (indispensable pour la scalabilité avec des fichiers volumineux).

    Note : fonctionne identiquement pour les .xlsx et les .csv — le hash porte
    sur les octets bruts du fichier, indépendamment du format.

    Args:
        path: Chemin du fichier à hacher.

    Returns:
        Empreinte hexadécimale SHA-256 (64 caractères).

    Raises:
        FileNotFoundError: Si le fichier n'existe pas.
        PermissionError: Si l'accès au fichier est refusé.
    """
    h = hashlib.new(SETTINGS.hash_algorithm)
    with open(path, "rb") as f:
        for block in iter(lambda: f.read(65_536), b""):
            h.update(block)
    return h.hexdigest()


def hash_bytes(data: bytes) -> str:
    """Calcule l'empreinte SHA-256 d'un contenu binaire en mémoire.

    Utilisé pour les fichiers uploadés via Streamlit (déjà chargés en RAM).
    Fonctionne pour tout format binaire : .xlsx, .csv, etc.

    Args:
        data: Contenu binaire du fichier.

    Returns:
        Empreinte hexadécimale SHA-256 (64 caractères).
    """
    return hashlib.new(SETTINGS.hash_algorithm, data).hexdigest()


# ---------------------------------------------------------------------------
# VALIDATION DE SCHÉMA
# ---------------------------------------------------------------------------

def validate_schema(
    df: pd.DataFrame,
    dictionary: Optional[dict[str, ColSpec]] = None,
) -> list[str]:
    """Identifie les colonnes requises absentes du DataFrame.

    Utilise ``DATA_DICTIONARY`` de ``config.py`` comme référence par défaut,
    ce qui garantit que toute évolution du schéma métier (ajout d'une colonne
    obligatoire) est automatiquement propagée sans modifier cette fonction.

    Args:
        df: DataFrame à valider. Les colonnes doivent être **déjà normalisées**
            (strip + lower) avant appel.
        dictionary: Dictionnaire de validation alternatif (utile pour les tests).
            Si None, utilise ``config.DATA_DICTIONARY``.

    Returns:
        Liste des noms de colonnes requises manquantes. Liste vide si tout est OK.
    """
    schema_dict = dictionary if dictionary is not None else DATA_DICTIONARY
    return [
        col_name
        for col_name, col_spec in schema_dict.items()
        if col_spec.is_required and col_name not in df.columns
    ]


def check_required_columns(
    df: pd.DataFrame,
    required: Optional[list[str]] = None,
) -> None:
    """Vérifie la présence des colonnes requises et lève ValueError si manquantes.

    Wrapper rétro-compatible : si ``required`` est fourni, utilise cette liste
    explicite (mode legacy) ; sinon délègue à ``validate_schema`` + ``DATA_DICTIONARY``
    (mode recommandé).

    Args:
        df: DataFrame à valider.
        required: Liste explicite de colonnes obligatoires (mode legacy).
            Si None, utilise ``DATA_DICTIONARY`` comme référence.

    Raises:
        ValueError: Si au moins une colonne requise est manquante.
            Le message indique les colonnes manquantes ET les colonnes attendues.
    """
    if required is not None:
        missing_cols = [c for c in required if c not in df.columns]
    else:
        missing_cols = validate_schema(df)

    if missing_cols:
        expected_cols = REQUIRED_COLUMNS
        raise ValueError(
            f"Colonnes requises manquantes : {missing_cols}. "
            f"Colonnes attendues après mapping : {expected_cols}. "
            f"Colonnes présentes dans le fichier : {list(df.columns)}"
        )


# ---------------------------------------------------------------------------
# RÉCONCILIATION - Checkpoint d'intégrité des données
# ---------------------------------------------------------------------------

@dataclass
class ReconciliationReport:
    """Résultat du contrôle d'intégrité entre données sources et données traitées.

    Ce rapport est la preuve mathématique que le pipeline de nettoyage n'a pas
    introduit de perte ou de modification silencieuse de valeur des données.

    Attributes:
        source_row_count: Nombre de lignes dans le DataFrame brut d'entrée.
        processed_row_count: Nombre de lignes après nettoyage complet.
        dropped_row_count: Différence (attendue = lignes à dates invalides).
        expected_source_sum: Somme des montants calculée depuis la source,
            uniquement sur les lignes dont la date est valide.
        processed_sum: Somme des montants après nettoyage (NaN exclus par pandas).
        absolute_gap: |expected_source_sum - processed_sum| en euros.
        gap_pct: Écart relatif en proportion de la somme source attendue.
        invalid_date_count: Lignes supprimées pour date non parseable.
        invalid_amount_count: Montants non numériques dans les lignes conservées
            (transformés en NaN, exclus du CA - tracés dans le log).
        integrity_ok: True si gap_pct ≤ SETTINGS.reconciliation_tolerance_pct.
        message: Message synthétique (✅ OK ou 🚨 ALERTE) pour les logs et l'UI.
    """

    source_row_count: int
    processed_row_count: int
    dropped_row_count: int
    expected_source_sum: float
    processed_sum: float
    absolute_gap: float
    gap_pct: float
    invalid_date_count: int
    invalid_amount_count: int
    integrity_ok: bool
    message: str


def reconcile_data(
    df_source: pd.DataFrame,
    df_processed: pd.DataFrame,
    date_col: str = "date",
    amount_col: str = "montant",
) -> ReconciliationReport:
    """Vérifie mathématiquement qu'aucune valeur des données n'a été perdue ou altérée.

    **Principe du Checkpoint :**
    La fonction recalcule indépendamment, depuis les données *sources brutes*,
    la somme que le DataFrame traité *devrait* contenir. Elle compare ensuite
    cette somme attendue à la somme réelle du DataFrame nettoyé.

    Cela garantit que le pipeline (``clean_data`` → ``add_month_column``) n'a
    introduit aucune transformation silencieuse autre que :
    1. La suppression des lignes à date invalide (comportement documenté).
    2. La mise à NaN des montants non numériques (comportement documenté).

    Toute autre divergence est une anomalie à investiguer avant diffusion.

    **Note de scalabilité (volumétrie ×100) :**
    Avec ~2M de lignes, cette fonction reste O(n) grâce à la vectorisation pandas.
    Au-delà de 10M de lignes, utiliser Polars (lazy evaluation) ou traiter par
    chunks avec ``pd.read_csv(..., chunksize=50_000)`` et agréger les sommes.

    Args:
        df_source: DataFrame brut **avant** tout nettoyage, colonnes normalisées
            (strip + lower) mais valeurs "sales" autorisées.
        df_processed: DataFrame **après** ``clean_data()`` et éventuellement
            ``add_month_column()``. La colonne ``montant`` est en float64.
        date_col: Nom de la colonne date dans df_source (après normalisation).
        amount_col: Nom de la colonne montant dans df_source et df_processed.

    Returns:
        ReconciliationReport contenant le verdict d'intégrité et tous les
        indicateurs pour le log d'audit.
    """
    source_count = len(df_source)
    processed_count = len(df_processed)

    # --- Simulation du nettoyage sur la source pour calculer la somme "attendue" ---
    df_sim = df_source.dropna(how="all").copy()

    # Étape A : normaliser les noms de colonnes (strip + lower)
    df_sim.columns = df_sim.columns.astype(str).str.strip().str.lower()

    # Étape B : fusionner les colonnes dupliquées
    if df_sim.columns.duplicated().any():
        for dup_col in df_sim.columns[df_sim.columns.duplicated()].unique():
            sub_df = df_sim.loc[:, df_sim.columns == dup_col]
            df_sim = df_sim.drop(columns=dup_col)
            df_sim[dup_col] = sub_df.bfill(axis=1).iloc[:, 0]

    # Étape 1 : Parser les dates
    if date_col in df_sim.columns:
        parsed_dates = pd.to_datetime(df_sim[date_col], dayfirst=True, errors="coerce")
    else:
        parsed_dates = pd.Series(
            [pd.NaT] * len(df_sim), index=df_sim.index, dtype="datetime64[ns]"
        )

    invalid_date_count = int(parsed_dates.isna().sum())
    valid_date_mask = parsed_dates.notna()

    # Étape 2 : Parser les montants sur les lignes à date valide seulement
    if amount_col in df_sim.columns and valid_date_mask.any():
        amounts_str = (
            df_sim.loc[valid_date_mask, amount_col]
            .astype(str)
            .str.strip()
            .str.replace(",", ".", regex=False)
            .str.replace("\u00a0", "", regex=False)
        )
        parsed_amounts = pd.to_numeric(amounts_str, errors="coerce")
        invalid_amount_count = int(parsed_amounts.isna().sum())
        expected_source_sum = float(parsed_amounts.sum())
    else:
        expected_source_sum = 0.0
        invalid_amount_count = 0

    # --- Somme réelle après traitement ---
    if amount_col in df_processed.columns:
        processed_sum = float(df_processed[amount_col].sum())
    else:
        processed_sum = 0.0

    # --- Calcul de l'écart ---
    absolute_gap = abs(expected_source_sum - processed_sum)

    if expected_source_sum != 0.0:
        gap_pct = absolute_gap / abs(expected_source_sum)
    else:
        gap_pct = 0.0 if processed_sum == 0.0 else 1.0

    tolerance = SETTINGS.reconciliation_tolerance_pct
    integrity_ok = gap_pct <= tolerance

    if integrity_ok:
        message = (
            f"✅ Intégrité OK - "
            f"écart {absolute_gap:.4f} € ({gap_pct * 100:.4f} %) "
            f"≤ tolérance {tolerance * 100:.3f} %"
        )
    else:
        message = (
            f"🚨 ALERTE INTÉGRITÉ - "
            f"écart {absolute_gap:.4f} € ({gap_pct * 100:.2f} %) "
            f"> tolérance {tolerance * 100:.3f} % - "
            f"Investigation requise avant toute diffusion du reporting."
        )

    logging.info(f"[Réconciliation] {message}")
    logging.info(
        f"[Réconciliation] Lignes : {source_count} source → {processed_count} traitées "
        f"({invalid_date_count} dates invalides, {invalid_amount_count} montants invalides)"
    )

    return ReconciliationReport(
        source_row_count=source_count,
        processed_row_count=processed_count,
        dropped_row_count=source_count - processed_count,
        expected_source_sum=expected_source_sum,
        processed_sum=processed_sum,
        absolute_gap=absolute_gap,
        gap_pct=gap_pct,
        invalid_date_count=invalid_date_count,
        invalid_amount_count=invalid_amount_count,
        integrity_ok=integrity_ok,
        message=message,
    )


# ---------------------------------------------------------------------------
# CHARGEMENT - Lecture CSV robuste (helper privé)
# ---------------------------------------------------------------------------

def _read_csv_robust(source: Any, filename: str) -> pd.DataFrame:
    """Lit un fichier CSV avec détection automatique du séparateur et fallback d'encodages.

    **Problématique CSV en contexte PME française :**
    Les exports de CRM, ERP et Excel français produisent des CSV hétérogènes :
    - Séparateur `;` (standard Excel FR) ou `,` (standard international)
    - Encodage latin-1/cp1252 (anciens systèmes Windows) ou utf-8 (modernes)
    - BOM utf-8-sig ajouté par Excel lors d'un export "CSV UTF-8"

    ``sep=None, engine="python"`` délègue la détection du séparateur au moteur
    Python de pandas (``csv.Sniffer``), qui inspecte les premières lignes du
    fichier. Plus coûteux que de fixer le séparateur manuellement, mais indispensable
    pour l'universalité (PME ≠ un seul système source).

    **Note de scalabilité :**
    Pour des CSV > 500 Mo, envisager ``pd.read_csv(..., chunksize=50_000)``
    et concaténer les chunks. La détection auto du séparateur ne supporte pas
    le mode chunked → pré-détecter le séparateur en lisant uniquement les
    premières lignes (``nrows=5``), puis relire en mode chunked.

    Args:
        source: Chemin disque (Path/str) ou buffer mémoire (io.BytesIO).
            En mode buffer (Streamlit), la position est réinitialisée (seek(0))
            avant chaque tentative d'encodage pour éviter une lecture partielle.
        filename: Nom du fichier (pour les messages de log uniquement).

    Returns:
        DataFrame pandas avec les données du CSV.

    Raises:
        ValueError: Si aucun encodage de la liste ``SETTINGS.csv_encodings_fallback``
            ne permet de lire le fichier correctement.
    """
    last_error: Exception = Exception("Aucune tentative effectuée")

    for encoding in SETTINGS.csv_encodings_fallback:
        try:
            # Réinitialisation du curseur à chaque tentative si c'est un buffer.
            # Indispensable : une tentative échouée déplace le curseur en position
            # arbitraire, ce qui corromprait la lecture suivante.
            if hasattr(source, "seek"):
                source.seek(0)

            df = pd.read_csv(
                source,
                sep=None,           # détection automatique du séparateur (Sniffer)
                engine="python",    # seul engine supportant sep=None
                encoding=encoding,
                on_bad_lines="warn",  # lignes malformées loggées, pas bloquantes
            )
            logging.info(
                f"CSV '{filename}' lu avec encodage '{encoding}' "
                f"({len(df)} lignes, {len(df.columns)} colonnes)"
            )
            return df

        except UnicodeDecodeError as exc:
            # Encodage incompatible : on tente le suivant dans la liste.
            logging.debug(
                f"CSV '{filename}' : encodage '{encoding}' incompatible → "
                f"tentative suivante. Détail : {exc}"
            )
            last_error = exc
        except Exception as exc:
            # Erreur structurelle (fichier vide, non-CSV, etc.) : inutile de
            # tester d'autres encodages, on abandonne immédiatement.
            logging.error(f"CSV '{filename}' : erreur structurelle → {exc}")
            raise ValueError(
                f"Impossible de lire le CSV '{filename}' : {exc}"
            ) from exc

    raise ValueError(
        f"Impossible de lire le CSV '{filename}' avec les encodages testés "
        f"({list(SETTINGS.csv_encodings_fallback)}). "
        f"Dernière erreur : {last_error}"
    )


def _read_single_file(source: Any, filename: str) -> pd.DataFrame:
    """Dispatcher : lit un fichier Excel ou CSV selon son extension.

    **Principe de conception (Single Responsibility) :**
    Ce helper centralise la décision de format en un seul endroit. Tous les
    appelants (``load_source_files``, ``read_uploaded_files``) délèguent ici,
    ce qui garantit qu'ajouter un nouveau format (ex: .ods) ne nécessite
    de modifier qu'une seule fonction.

    Les deux branches produisent un ``pd.DataFrame`` identique en sortie.
    Tout le reste du pipeline (nettoyage, réconciliation, export) est
    **format-agnostique** et ne nécessite aucune modification.

    Args:
        source: Chemin disque (Path/str) ou buffer mémoire (io.BytesIO).
        filename: Nom du fichier, utilisé uniquement pour détecter l'extension
            et pour les messages de log.

    Returns:
        DataFrame pandas brut (colonnes non normalisées, valeurs brutes).

    Raises:
        ValueError: Si le format n'est pas reconnu ou si la lecture échoue.
    """
    ext = Path(filename).suffix.lower().lstrip(".")

    if ext == "csv":
        return _read_csv_robust(source, filename)
    elif ext == "xlsx":
        return pd.read_excel(source, engine="openpyxl")
    else:
        raise ValueError(
            f"Format non supporté : '.{ext}' (fichier : '{filename}'). "
            f"Formats acceptés : {sorted(_SOURCE_EXTENSIONS)}."
        )


# ---------------------------------------------------------------------------
# CHARGEMENT - Mode Batch (disque)
# ---------------------------------------------------------------------------

def load_source_files(folder_path: str) -> pd.DataFrame:
    """Lit tous les fichiers sources (.xlsx et .csv) d'un répertoire et les concatène.

    Continue même si certains fichiers sont corrompus, protégés ou mal formés.
    Chaque fichier lisible est tracé avec son hash SHA-256 pour l'audit.

    Les fichiers temporaires Excel (``~$...``) sont ignorés automatiquement
    (Excel les crée quand un fichier est ouvert par un autre processus).

    **Extensions supportées :** définies dans ``SETTINGS.csv_source_extensions``
    (par défaut : xlsx, csv). Modifiable sans toucher au code.

    **Note de scalabilité (volumétrie ×100) :**
    Avec 10M+ de lignes, privilégier :
    - ``pd.read_csv(..., engine="pyarrow")`` pour les CSV (x3-5 plus rapide) ;
    - Migration vers Parquet/DuckDB pour le stockage intermédiaire ;
    - Polars ``pl.read_csv`` / ``pl.read_excel`` avec lazy evaluation.
    Complexité globale : O(n × k) avec n = lignes totales, k = fichiers.

    Args:
        folder_path: Chemin du répertoire contenant les fichiers sources.

    Returns:
        DataFrame consolidé (toutes lignes, toutes colonnes).

    Raises:
        ValueError: Si aucun fichier n'est lisible dans le répertoire.
    """
    folder = Path(folder_path)

    # Collecte de tous les fichiers des formats supportés, en excluant les
    # fichiers temporaires Excel (préfixe ~$) qui ne sont pas des données réelles.
    files: list[Path] = []
    for ext in SETTINGS.csv_source_extensions:
        files.extend(
            f for f in folder.glob(f"*.{ext}")
            if not f.name.startswith("~$")
        )

    # Tri alphabétique → ordre d'exécution reproductible d'une run à l'autre.
    files = sorted(files)

    logging.info(
        f"{len(files)} fichier(s) source(s) détecté(s) dans '{folder_path}' "
        f"(formats : {sorted(_SOURCE_EXTENSIONS)})"
    )

    dfs: list[pd.DataFrame] = []
    error_count = 0

    for file_path in files:
        try:
            file_hash = hash_file_on_disk(file_path)
            fmt = file_path.suffix.upper().lstrip(".")
            logging.info(
                f"Lecture [{fmt}] : {file_path.name} | "
                f"{SETTINGS.hash_algorithm.upper()} : {file_hash}"
            )
            df = _read_single_file(str(file_path), file_path.name)
            dfs.append(df)
            logging.info(f"  ✓ {len(df)} lignes lues")

        except Exception as exc:
            error_count += 1
            logging.error(f"  ✗ Échec lecture : {file_path.name} | {exc}")

    if error_count:
        logging.warning(f"{error_count} fichier(s) ignoré(s) pour erreur de lecture.")

    if not dfs:
        raise ValueError(
            f"Aucun fichier source lisible dans '{folder_path}'. "
            "Vérifiez que le dossier contient des fichiers .xlsx ou .csv "
            "valides et non protégés."
        )

    return pd.concat(dfs, ignore_index=True)


# Alias de rétrocompatibilité : main.py v1 appelait load_excel_files().
# Permet une migration progressive sans casser les scripts existants.
load_excel_files = load_source_files


# ---------------------------------------------------------------------------
# CHARGEMENT - Mode Streamlit (fichiers uploadés en mémoire)
# ---------------------------------------------------------------------------

def read_uploaded_files(uploaded_files: list[Any]) -> LoadResult:
    """Lit des fichiers uploadés via Streamlit (.xlsx et .csv) et les concatène.

    Accepte tout objet ayant un attribut ``name`` et compatible avec
    ``pd.read_excel()`` ou ``pd.read_csv()`` (ex: ``streamlit.UploadedFile``,
    ``io.BytesIO``). Continue même si certains fichiers sont illisibles.

    **Détection du format :** basée sur l'extension du nom de fichier (attribut
    ``name``), pas sur le contenu (magic bytes). Justification : les PME renomment
    rarement leurs exports, et la détection par contenu serait plus complexe pour
    un gain marginal.

    Args:
        uploaded_files: Liste de fichiers uploadés (objets Streamlit ou BytesIO).

    Returns:
        Tuple (df_concatené, fichiers_en_erreur).
        Les fichiers en erreur sont listés sous forme (nom, message_erreur).

    Raises:
        ValueError: Si aucun fichier n'est lisible.
    """
    dfs: list[pd.DataFrame] = []
    failed_files: FailedFiles = []

    for f in uploaded_files:
        filename = getattr(f, "name", "fichier_sans_nom.xlsx")
        ext = Path(filename).suffix.lower().lstrip(".")

        if ext not in _SOURCE_EXTENSIONS:
            failed_files.append((
                filename,
                f"Extension '.{ext}' non supportée. Formats acceptés : "
                f"{sorted(_SOURCE_EXTENSIONS)}."
            ))
            logging.warning(f"Extension non supportée ignorée : {filename}")
            continue

        try:
            # Pour les BytesIO uploadés via Streamlit, on crée un buffer frais
            # afin de permettre une re-lecture propre (getvalue() retourne
            # les bytes depuis le début, indépendamment de la position du curseur).
            if hasattr(f, "getvalue"):
                source = io.BytesIO(f.getvalue())
            else:
                source = f

            df = _read_single_file(source, filename)
            dfs.append(df)
            fmt = ext.upper()
            logging.info(f"Upload OK [{fmt}] : {filename} ({len(df)} lignes)")

        except Exception as exc:
            failed_files.append((filename, str(exc)))
            logging.error(f"Upload KO : {filename} | {exc}")

    if not dfs:
        raise ValueError(
            "Aucun fichier source lisible parmi ceux uploadés. "
            "Vérifiez le format (doit être .xlsx ou .csv réel, "
            "non protégé, non corrompu)."
        )

    return pd.concat(dfs, ignore_index=True), failed_files


def read_one_uploaded_file(f: Any) -> pd.DataFrame:
    """Lit un seul fichier uploadé et retourne son DataFrame brut non nettoyé.

    Utilisé par app.py pour la déduplication inter-formats (ex: même fichier
    présent en .xlsx et en .csv). Lire chaque fichier individuellement avant
    de les concaténer permet de calculer une empreinte *métier* par fichier.

    Args:
        f: Fichier uploadé Streamlit (UploadedFile) ou io.BytesIO.
            Doit avoir un attribut ``name`` pour la détection du format.

    Returns:
        DataFrame brut (colonnes et valeurs telles que lues, sans nettoyage).

    Raises:
        ValueError: Si le format n'est pas supporté ou si la lecture échoue.
    """
    filename = getattr(f, "name", "fichier_sans_nom")
    # Créer un buffer frais depuis getvalue() pour garantir une lecture depuis
    # le début, indépendamment de la position courante du curseur de f.
    if hasattr(f, "getvalue"):
        source = io.BytesIO(f.getvalue())
    else:
        source = f
    return _read_single_file(source, filename)


# Alias de rétrocompatibilité : app.py v1 appelait read_uploaded_excels().
read_uploaded_excels = read_uploaded_files


# ---------------------------------------------------------------------------
# NETTOYAGE
# ---------------------------------------------------------------------------

def clean_data(df: pd.DataFrame) -> pd.DataFrame:
    """Normalise et nettoie un DataFrame brut pour le rendre exploitable.

    Transformations appliquées (dans l'ordre) :
    1. **Normalisation des noms de colonnes** : strip + lower.
       Raison : les exports Excel/CSV ont souvent des espaces parasites et des casses variées.
    2. **Fusion des colonnes dupliquées** : priorité à la première valeur non-nulle.
       Raison : certains exports concatenent des fichiers avec des colonnes
       portant le même nom après normalisation.
    3. **Suppression des lignes entièrement vides** : artefacts Excel/CSV courants.
    4. **Validation du schéma** via ``DATA_DICTIONARY`` (colonnes requises).
    5. **Nettoyage des montants** : gestion des formats FR (virgule) et EN (point),
       espaces insécables, valeurs texte → NaN conservés (tracés, exclus du CA).
    6. **Parsing des dates** : ``dayfirst=True`` pour les formats JJ/MM/AAAA.
       Les lignes à date invalide sont **supprimées** (règle comptable : une
       transaction sans date n'est pas imputable à une période).
    7. **Normalisation du champ commercial** en Title Case.

    **Note de scalabilité :**
    Toutes les opérations utilisent la vectorisation pandas (C-level). Sur 2M
    de lignes, le goulot serait ``pd.to_datetime`` (parsing regex interne).
    Alternative : pré-formater les dates en ISO dans les sources, ou utiliser
    Polars ``pl.col("date").str.strptime()``.

    Args:
        df: DataFrame brut après chargement. Les colonnes peuvent être "sales"
            (espaces, majuscules, virgules dans les montants, dates invalides).
            Fonctionne identiquement qu'il provienne d'un .xlsx ou d'un .csv.

    Returns:
        DataFrame nettoyé avec colonnes ``date`` (datetime64) et ``montant`` (float64).

    Raises:
        ValueError: Si les colonnes requises sont absentes après normalisation.
    """
    df = df.copy()  # ne pas muter l'entrée (principe d'immuabilité fonctionnelle)

    # --- 1. Normalisation des noms de colonnes ---
    df.columns = df.columns.astype(str).str.strip().str.lower()

    # --- 2. Fusion des colonnes dupliquées ---
    if df.columns.duplicated().any():
        duplicate_names: list[str] = df.columns[df.columns.duplicated()].unique().tolist()
        logging.warning(f"Colonnes dupliquées détectées (fusion) : {duplicate_names}")
        for col_name in duplicate_names:
            sub_df = df.loc[:, df.columns == col_name]
            merged_value = sub_df.bfill(axis=1).iloc[:, 0]
            df = df.drop(columns=col_name)
            df[col_name] = merged_value

    # --- 3. Suppression lignes entièrement vides ---
    count_before_dropna = len(df)
    df = df.dropna(how="all")
    empty_row_count = count_before_dropna - len(df)
    if empty_row_count:
        logging.info(f"Lignes entièrement vides supprimées : {empty_row_count}")

    # --- 4. Validation du schéma via DATA_DICTIONARY ---
    check_required_columns(df)

    # --- 5. Nettoyage et conversion des montants ---
    raw_amounts = df["montant"].copy()

    amounts_str = (
        df["montant"]
        .astype(str)
        .str.strip()
        .str.replace(",", ".", regex=False)
        .str.replace("\u00a0", "", regex=False)
        .str.replace(" ", "", regex=False)
    )
    df["montant"] = pd.to_numeric(amounts_str, errors="coerce")

    empty_source_count = int(
        raw_amounts.isna().sum()
        + (raw_amounts.astype(str).str.strip() == "").sum()
    )
    if empty_source_count:
        logging.info(f"Montants vides en source : {empty_source_count}")

    invalid_mask = (
        df["montant"].isna()
        & raw_amounts.notna()
        & (raw_amounts.astype(str).str.strip() != "")
    )
    invalid_count = int(invalid_mask.sum())
    if invalid_count:
        examples = raw_amounts.loc[invalid_mask].head(5).tolist()
        logging.warning(
            f"Montants non convertibles → NaN : {invalid_count} (exclus du CA). "
            f"Exemples : {examples}"
        )

    # --- 6. Parsing et filtrage des dates ---
    df["date"] = pd.to_datetime(df["date"], dayfirst=True, errors="coerce")
    nat_count = int(df["date"].isna().sum())
    if nat_count:
        logging.warning(f"Dates invalides → supprimées : {nat_count} lignes")

    count_before_date_filter = len(df)
    df = df.dropna(subset=["date"])
    dropped_count = count_before_date_filter - len(df)
    logging.info(f"Lignes supprimées (date invalide) : {dropped_count} | Restantes : {len(df)}")

    # --- 7. Normalisation champ commercial ---
    if "commercial" in df.columns:
        df["commercial"] = df["commercial"].astype(str).str.strip().str.title()

    return df.reset_index(drop=True)


# ---------------------------------------------------------------------------
# ENRICHISSEMENT
# ---------------------------------------------------------------------------

def add_month_column(df: pd.DataFrame) -> pd.DataFrame:
    """Ajoute une colonne ``mois`` au format ``YYYY-MM`` (string).

    La colonne ``mois`` est la clé d'agrégation pour les reportings mensuels.
    Le format ``YYYY-MM`` est choisi pour :
    - Tri lexicographique = tri chronologique (pas besoin de conversion).
    - Interopérabilité (Excel, SQL, JSON).

    Args:
        df: DataFrame avec une colonne ``date`` de type datetime64.

    Returns:
        DataFrame enrichi d'une colonne ``mois`` (str, ex: "2025-01").
    """
    df = df.copy()
    df["mois"] = df["date"].dt.to_period("M").astype(str)
    return df


# ---------------------------------------------------------------------------
# REPORTINGS
# ---------------------------------------------------------------------------

def aggregate_by_month(df: pd.DataFrame) -> pd.DataFrame:
    """Agrège le chiffre d'affaires par mois.

    ``skipna=True`` (comportement par défaut de sum()) → les montants NaN
    sont exclus du CA, conformément au traitement comptable attendu.

    Args:
        df: DataFrame nettoyé avec colonnes ``mois`` et ``montant``.

    Returns:
        DataFrame avec colonnes [mois, montant], trié chronologiquement.
    """
    return (
        df.groupby("mois", sort=True)["montant"]
        .sum()
        .reset_index()
    )


def aggregate_by_salesperson(df: pd.DataFrame) -> pd.DataFrame:
    """Agrège le chiffre d'affaires par commercial.

    Retourne un DataFrame vide (avec les bonnes colonnes) si la colonne
    ``commercial`` est absente, évitant toute exception dans l'UI.

    Args:
        df: DataFrame nettoyé, avec ou sans colonne ``commercial``.

    Returns:
        DataFrame avec colonnes [commercial, montant], trié par montant décroissant.
        DataFrame vide si colonne ``commercial`` absente.
    """
    if "commercial" not in df.columns:
        return pd.DataFrame(columns=["commercial", "montant"])

    return (
        df.groupby("commercial")["montant"]
        .sum()
        .sort_values(ascending=False)
        .reset_index()
    )


# ---------------------------------------------------------------------------
# EXPORTS - Helpers privés de mise en forme Excel
# ---------------------------------------------------------------------------

def _format_excel_sheet(ws: Any) -> None:
    """Applique le formatage openpyxl standard à une feuille Excel."""
    from openpyxl.styles import Font
    from openpyxl.styles import numbers as openpyxl_numbers

    for cell in ws[1]:
        cell.font = Font(bold=True)

    ws.freeze_panes = "A2"

    for col_cells in ws.columns:
        col_letter = col_cells[0].column_letter
        header = str(col_cells[0].value) if col_cells[0].value else ""

        max_len = len(header)
        for cell in col_cells[1:]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))

        width = max_len + 4
        width = max(width, SETTINGS.excel_min_col_width)

        if any(kw in header.lower() for kw in ("montant", "euro", "total", "ca")):
            width = max(width, SETTINGS.excel_money_min_width)

        ws.column_dimensions[col_letter].width = min(width, SETTINGS.excel_max_col_width)

    headers = [cell.value for cell in ws[1]]
    header_idx: dict[str, int] = {h: i + 1 for i, h in enumerate(headers) if h}

    if "date" in header_idx:
        c = header_idx["date"]
        for r in range(2, ws.max_row + 1):
            ws.cell(r, c).number_format = "DD/MM/YYYY"

    for monetary_col in ("montant", "montant en euros", "total"):
        if monetary_col in header_idx:
            c = header_idx[monetary_col]
            for r in range(2, ws.max_row + 1):
                ws.cell(r, c).number_format = openpyxl_numbers.FORMAT_CURRENCY_EUR_SIMPLE


def _prepare_report_exports(
    report_months: pd.DataFrame,
    report_salespeople: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Prépare des copies des rapports pour l'export (renommage colonne montant)."""
    months_copy = report_months.copy()
    salespeople_copy = report_salespeople.copy()

    if "montant" in months_copy.columns:
        months_copy = months_copy.rename(columns={"montant": "montant en euros"})
    if "montant" in salespeople_copy.columns:
        salespeople_copy = salespeople_copy.rename(columns={"montant": "montant en euros"})

    return months_copy, salespeople_copy


# ---------------------------------------------------------------------------
# EXPORTS - Mode Batch (fichiers sur disque)
# ---------------------------------------------------------------------------

def export_excel(
    df: pd.DataFrame,
    report_months: pd.DataFrame,
    report_salespeople: pd.DataFrame,
    filepath: Path,
) -> Path:
    """Exporte les données et rapports dans un fichier Excel multi-feuilles."""
    Path(filepath).parent.mkdir(parents=True, exist_ok=True)
    months_export, salespeople_export = _prepare_report_exports(
        report_months, report_salespeople
    )

    try:
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Donnees_brutes", index=False)
            months_export.to_excel(writer, sheet_name="Total_par_mois", index=False)
            salespeople_export.to_excel(
                writer, sheet_name="Total_par_commercial", index=False
            )

        logging.info(f"Excel exporté : {filepath}")
        return Path(filepath)

    except Exception as exc:
        logging.error(f"Erreur export Excel : {filepath} | {exc}")
        raise


def format_excel_file(filepath: Path) -> Path:
    """Applique le formatage openpyxl à un fichier Excel existant sur le disque."""
    from openpyxl import load_workbook

    wb = load_workbook(filepath)
    for ws in wb.worksheets:
        _format_excel_sheet(ws)

    wb.save(filepath)
    logging.info(f"Excel formaté : {filepath}")
    return filepath


# ---------------------------------------------------------------------------
# EXPORTS - Mode Streamlit (bytes en mémoire)
# ---------------------------------------------------------------------------

# Géré via workbook.py

# ---------------------------------------------------------------------------
# EXPORTS - PDF
# ---------------------------------------------------------------------------

def _build_pdf_elements(
    report_months: pd.DataFrame,
    report_salespeople: pd.DataFrame,
) -> list[Any]:
    """Construit la liste des éléments ReportLab pour le PDF."""
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, Spacer, Table, TableStyle

    styles = getSampleStyleSheet()
    elements: list[Any] = []

    min_month = report_months["mois"].min() if not report_months.empty else "N/A"
    year = min_month.split("-")[0] if min_month != "N/A" else "N/A"

    elements.append(
        Paragraph(f"Rapport - {SETTINGS.app_name} {year}", styles["Heading1"])
    )
    elements.append(
        Paragraph(f"Généré le {datetime.today().strftime('%d/%m/%Y')}", styles["Normal"])
    )
    elements.append(
        Paragraph(f"Version : {SETTINGS.app_version}", styles["Normal"])
    )
    elements.append(Spacer(1, 12))

    report_months = report_months.copy()
    report_salespeople = report_salespeople.copy()
    for d in (report_months, report_salespeople):
        if "montant" in d.columns:
            d["montant"] = pd.to_numeric(d["montant"], errors="coerce").round(2)
            d.rename(columns={"montant": "montant en euros"}, inplace=True)

    def _add_table(table_df: pd.DataFrame, title: str) -> None:
        elements.append(Paragraph(title, styles["Heading2"]))
        elements.append(Spacer(1, 6))

        if table_df is None or table_df.empty:
            elements.append(Paragraph("Aucune donnée disponible.", styles["Normal"]))
            elements.append(Spacer(1, 12))
            return

        data = [list(table_df.columns)] + table_df.values.tolist()
        table = Table(data, hAlign="LEFT")
        table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("PADDING", (0, 0), (-1, -1), 6),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
            ])
        )
        elements.append(table)
        elements.append(Spacer(1, 12))

    _add_table(report_months, "Total par mois")
    _add_table(report_salespeople, "Total par commercial")

    return elements


def build_pdf_bytes(
    report_months: pd.DataFrame,
    report_salespeople: pd.DataFrame,
) -> bytes:
    """Génère un rapport PDF en mémoire (pour téléchargement Streamlit)."""
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    doc.build(_build_pdf_elements(report_months, report_salespeople))
    buffer.seek(0)
    return buffer.getvalue()


def export_pdf(
    report_months: pd.DataFrame,
    report_salespeople: pd.DataFrame,
    out_dir: Path,
) -> Path:
    """Exporte le rapport PDF sur le disque."""
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate

    out_dir.mkdir(parents=True, exist_ok=True)

    min_month = report_months["mois"].min()
    max_month = report_months["mois"].max()
    filename = (
        f"reporting_{min_month}.pdf"
        if min_month == max_month
        else f"reporting_{min_month}_to_{max_month}.pdf"
    )

    pdf_path = out_dir / filename
    doc = SimpleDocTemplate(str(pdf_path), pagesize=A4)
    doc.build(_build_pdf_elements(report_months, report_salespeople))

    logging.info(f"PDF généré : {pdf_path}")
    return pdf_path